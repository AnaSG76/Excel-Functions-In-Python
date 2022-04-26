from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# Data set example: Iowa State University undergraduate tuition over the past ten years.
# Credits to techwithtim for some of the source code github repo url:
# https://github.com/techwithtim/ExcelPythonTutorial/blob/main/working.py

# data set
data = {
    "2012-13": {
        "Undergraduate_In_State": 7726,
        "Undergraduate_Out_of_State": 19838,
    },
    "2013-14": {
        "Undergraduate_In_State": 7726,
        "Undergraduate_Out_of_State": 20278,
    },
    "2014-15": {
        "Undergraduate_In_State": 7731,
        "Undergraduate_Out_of_State": 20617,
    },
    "2015-16": {
        "Undergraduate_In_State": 7736,
        "Undergraduate_Out_of_State": 20856,
    },
    "2016-17": {
        "Undergraduate_In_State": 8219,
        "Undergraduate_Out_of_State": 21583,
    },
    "2017-18": {
        "Undergraduate_In_State": 8636,
        "Undergraduate_Out_of_State": 22472,
    },
    "2018-19": {
        "Undergraduate_In_State": 8988,
        "Undergraduate_Out_of_State": 23392,
    },
    "2019-20": {
        "Undergraduate_In-State": 9320,
        "Undergraduate_Out_of_State": 24508,
    },
    "2020-21": {
        "Undergraduate_In-State": 9316,
        "Undergraduate_Out_of_State": 24504,
    },
    "2021-22": {
        "Undergraduate_In-State": 9634,
        "Undergraduate_Out-of-State": 25446,
    }
}

# write data into excel
wb = Workbook()
ws = wb.active
ws.title = "ISU_Tuition"

# set the width of the column
ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 26

headings = ['Academic_Year'] + list(data['2012-13'].keys())
ws.append(headings)

for academic_year in data:
    tuition = list(data[academic_year].values())
    ws.append([academic_year] + tuition)

# add some styling to header column and data in the table
for col in range(1, 4):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="00FF0000")
    ws[get_column_letter(col) +
       '1'].fill = PatternFill("solid", start_color="FFD700")

for rows in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    for cell in rows:
        cell.fill = PatternFill(
            start_color='00FF0000',
            end_color='00FF0000',
            fill_type="solid")
        cell.font = Font(bold=False, color="FFD700")

# save workbook
wb.save("ISU_Tuition_10Yrs.xlsx")

# Credits to College Tuition Compare website for the below data set.
# URL:
# https://www.collegetuitioncompare.com/trends/iowa-state-university/cost-of-attendance/#:~:text=The%20tuition%20%26%20fees%20have%20increased,to%20%2426%2C948%20(year%202022)
